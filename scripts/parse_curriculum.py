import openpyxl, json, re, unicodedata

wb = openpyxl.load_workbook('/mnt/user-data/uploads/tantervi_halo_2026-2027_uzemmernok.xlsx', data_only=True)
ws = wb['Tanterv']
data = [[ws.cell(r,c).value for c in range(1,7)] for r in range(1, ws.max_row+1)]

def slugify(s):
    s = s.strip().lower()
    s = s.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ö','o').replace('ő','o').replace('ú','u').replace('ü','u').replace('ű','u')
    s = re.sub(r'[^a-z0-9]+','_', s)
    return s.strip('_')

# note markers => text that is a note, not a prereq
NOTE_KEYS = ["ea és gyak","e és gyak","függetlenül","kritikus","kiugrási","téma:","160 óra",
             "java","c++","c programozás","javascript","előzetes teljesítése vagy párhuzamos",
             "egymástól függetlenül","az előadás","elismerhet","párhuzamos","felvehető"]

def split_field(text):
    """Return (prereq_str, note_str). Split on / then classify each chunk."""
    if not text: return "", ""
    chunks = re.split(r'\s*/\s*', text.replace("\n"," / "))
    prq, notes = [], []
    for c in chunks:
        c=c.strip()
        if not c: continue
        low=c.lower()
        if any(k in low for k in NOTE_KEYS):
            notes.append(c)
        else:
            prq.append(c)
    return " / ".join(prq), " / ".join(notes)

def prereq_names(prereq_str):
    """Extract candidate subject names from a prereq string (best-effort)."""
    if not prereq_str: return []
    # split on ; and , and VAGY
    parts = re.split(r'\s*;\s*|\s*,\s*|\s+VAGY\s+', prereq_str)
    out=[]
    for p in parts:
        p=p.strip()
        # strip trailing gy / ea qualifiers and periods
        p=re.sub(r'\s+(gy|ea|e)\.?$','',p, flags=re.I).strip(' .')
        if p and len(p)>2:
            out.append(p)
    return out

subjects=[]
sem=None; sem_num=None; group=None; sem_label=None
for row in data:
    a=row[0]
    if a is None: continue
    a=str(a).strip()
    if a.startswith("Üzemmérnök") or a.startswith("6 félév"): continue
    if a=="Tárgy neve": continue
    m=re.match(r'^(\d+)\.\s*félév', a)
    if m:
        sem_num=int(m.group(1)); sem_label=a; group=None; continue
    if a.startswith("Választható tárgyak, félév") or a.startswith("Választható tárgyak, félév nincs"):
        sem_num=None; sem_label=a; group="valaszthato"; continue
    if a in ("Kötelező tárgyak","Választható tárgyak"):
        group = "kotelezo" if a.startswith("Kötelező") else "valaszthato"
        continue
    # subject row
    ea_kr=row[1]; gy_kr=row[2]; ea_h=row[3]; gy_h=row[4]
    prq,note=split_field(row[5])
    ea_kr_v = ea_kr if isinstance(ea_kr,(int,float)) else 0
    gy_kr_v = gy_kr if isinstance(gy_kr,(int,float)) else 0
    subjects.append({
        "id": slugify(a),
        "name": a,
        "semester": sem_num,
        "semesterLabel": sem_label,
        "type": group or "kotelezo",
        "credits": {"ea": ea_kr, "gy": gy_kr, "total": (ea_kr_v+gy_kr_v)},
        "hours": {"ea": ea_h, "gy": gy_h},
        "prereqRaw": prq,
        "prereqNames": prereq_names(prq),
        "note": note,
        "completed": False,
    })

# resolve prereq names -> ids where possible
name_to_id = {s["name"].lower(): s["id"] for s in subjects}
# also index without trailing roman/qualifier noise
for s in subjects:
    ids=[]
    for nm in s["prereqNames"]:
        key=nm.lower()
        if key in name_to_id:
            ids.append(name_to_id[key])
        else:
            # fuzzy: startswith match
            cand=[sid for n,sid in name_to_id.items() if n.startswith(key[:12])]
            if len(cand)==1: ids.append(cand[0])
    s["prereqIds"]=sorted(set(ids))

print(f"parsed {len(subjects)} subjects")
with open('/home/claude/curriculum.json','w',encoding='utf-8') as f:
    json.dump(subjects,f,ensure_ascii=False,indent=1)
# quick sanity
resolved=sum(1 for s in subjects if s["prereqIds"])
print(f"{resolved} subjects have resolved prereq ids")
